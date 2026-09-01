# exporter.py
"""
Export schedule data to various formats (Excel, CSV, PDF).
"""

import csv
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF

from database import db_manager
from models import ScheduleEntry


class ScheduleExporter:
    """Handles export of schedule data to various formats."""
    
    def __init__(self, export_dir: str = "exports"):
        self.export_dir = Path(export_dir)
        self.export_dir.mkdir(exist_ok=True)
    
    def _get_schedule_data(self, filter_criteria: Optional[dict] = None) -> list:
        """Fetch schedule data from database with optional filters."""
        query = """
            SELECT 
                s.id,
                s.course_id,
                c.title as course_title,
                s.lecture_number,
                s.teacher_id,
                t.full_name as teacher_name,
                s.room_id,
                r.room_name,
                r.room_type,
                s.day,
                s.start_period,
                s.duration
            FROM schedule s
            JOIN courses c ON s.course_id = c.id
            JOIN teachers t ON s.teacher_id = t.id
            JOIN rooms r ON s.room_id = r.id
            WHERE 1=1
        """
        params = []
        
        if filter_criteria:
            if 'teacher_id' in filter_criteria:
                query += " AND s.teacher_id = ?"
                params.append(filter_criteria['teacher_id'])
            if 'room_id' in filter_criteria:
                query += " AND s.room_id = ?"
                params.append(filter_criteria['room_id'])
            if 'course_id' in filter_criteria:
                query += " AND s.course_id = ?"
                params.append(filter_criteria['course_id'])
            if 'day' in filter_criteria:
                query += " AND s.day = ?"
                params.append(filter_criteria['day'])
        
        query += " ORDER BY s.day, s.start_period, r.room_name"
        
        with db_manager.get_connection() as conn:
            cursor = conn.execute(query, params)
            return [dict(row) for row in cursor.fetchall()]
    
    def export_csv(self, filename: Optional[str] = None, **filters) -> str:
        """Export schedule to CSV format."""
        if filename is None:
            filename = f"schedule_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        filepath = self.export_dir / filename
        data = self._get_schedule_data(filters)
        
        if not data:
            raise ValueError("No schedule data to export")
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
        
        return str(filepath)
    
    def export_excel(self, filename: Optional[str] = None, **filters) -> str:
        """Export schedule to formatted Excel file."""
        if filename is None:
            filename = f"schedule_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        filepath = self.export_dir / filename
        data = self._get_schedule_data(filters)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Schedule"
        
        # Styles
        header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        cell_alignment = Alignment(vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['Day', 'Start Period', 'End Period', 'Course', 'Lecture #', 
                   'Teacher', 'Room', 'Room Type']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data
        for row_idx, entry in enumerate(data, 2):
            ws.cell(row=row_idx, column=1, value=entry['day']).border = thin_border
            ws.cell(row=row_idx, column=2, value=entry['start_period']).border = thin_border
            ws.cell(row=row_idx, column=3, 
                   value=entry['start_period'] + entry['duration'] - 1).border = thin_border
            ws.cell(row=row_idx, column=4, value=entry['course_title']).border = thin_border
            ws.cell(row=row_idx, column=5, value=entry['lecture_number'] + 1).border = thin_border
            ws.cell(row=row_idx, column=6, value=entry['teacher_name']).border = thin_border
            ws.cell(row=row_idx, column=7, value=entry['room_name']).border = thin_border
            ws.cell(row=row_idx, column=8, value=entry['room_type']).border = thin_border
            
            for col in range(1, 9):
                ws.cell(row=row_idx, column=col).alignment = cell_alignment
        
        # Adjust column widths
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Add summary sheet
        self._add_summary_sheet(wb, data)
        
        wb.save(filepath)
        return str(filepath)
    
    def _add_summary_sheet(self, wb, data):
        """Add a summary sheet with statistics."""
        ws = wb.create_sheet("Summary")
        
        # Calculate summary statistics
        total_lectures = len(data)
        unique_courses = len(set(d['course_id'] for d in data))
        unique_teachers = len(set(d['teacher_id'] for d in data))
        unique_rooms = len(set(d['room_id'] for d in data))
        
        summary_data = [
            ("Total Lectures Scheduled", total_lectures),
            ("Unique Courses", unique_courses),
            ("Unique Teachers", unique_teachers),
            ("Unique Rooms Used", unique_rooms),
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, 1):
            ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=value)
    
    def export_pdf(self, filename: Optional[str] = None, **filters) -> str:
        """Export schedule to PDF format."""
        if filename is None:
            filename = f"schedule_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        filepath = self.export_dir / filename
        data = self._get_schedule_data(filters)
        
        pdf = FPDF(orientation='L', unit='mm', format='A4')
        pdf.add_page()
        pdf.set_font('Arial', 'B', 16)
        pdf.cell(0, 10, 'University Schedule', 0, 1, 'C')
        pdf.ln(5)
        
        # Table headers
        pdf.set_font('Arial', 'B', 9)
        col_widths = [15, 20, 20, 55, 12, 40, 25, 25]
        headers = ['Day', 'Start', 'End', 'Course', 'Lec#', 'Teacher', 'Room', 'Type']
        
        for i, (header, width) in enumerate(zip(headers, col_widths)):
            pdf.cell(width, 7, header, 1)
        pdf.ln()
        
        # Table data
        pdf.set_font('Arial', '', 8)
        for entry in data:
            pdf.cell(col_widths[0], 6, str(entry['day']), 1)
            pdf.cell(col_widths[1], 6, str(entry['start_period']), 1)
            pdf.cell(col_widths[2], 6, str(entry['start_period'] + entry['duration'] - 1), 1)
            pdf.cell(col_widths[3], 6, entry['course_title'][:30], 1)
            pdf.cell(col_widths[4], 6, str(entry['lecture_number'] + 1), 1)
            pdf.cell(col_widths[5], 6, entry['teacher_name'][:25], 1)
            pdf.cell(col_widths[6], 6, entry['room_name'][:15], 1)
            pdf.cell(col_widths[7], 6, entry['room_type'][:12], 1)
            pdf.ln()
        
        pdf.output(filepath)
        return str(filepath)